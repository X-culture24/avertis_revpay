"""
Excel Export Service for Invoice Data
Provides functionality to export invoice data to Excel format
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from io import BytesIO


class ExcelExportService:
    """Service for exporting invoice data to Excel format"""
    
    def export_invoices(self, invoices, date_range=None):
        """
        Export invoices to Excel format
        
        Args:
            invoices: QuerySet of Invoice objects
            date_range: Optional dict with 'start' and 'end' dates
            
        Returns:
            BytesIO object containing Excel file
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Invoices"
        
        # Headers
        headers = [
            "Invoice Number",
            "Date",
            "Customer Name",
            "Amount (KES)",
            "Status",
            "Receipt Number"
        ]
        
        # Apply header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Filter by date range if provided
        if date_range:
            invoices = invoices.filter(
                created_at__gte=date_range['start'],
                created_at__lte=date_range['end']
            )
        
        # Data rows
        for row, invoice in enumerate(invoices, start=2):
            worksheet.cell(row=row, column=1).value = invoice.invoice_no
            worksheet.cell(row=row, column=2).value = invoice.created_at.strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row, column=3).value = invoice.customer_name
            worksheet.cell(row=row, column=4).value = float(invoice.total_amount)
            worksheet.cell(row=row, column=5).value = invoice.status.upper()
            worksheet.cell(row=row, column=6).value = invoice.receipt_no or 'N/A'
            
            # Center align all cells
            for col in range(1, 7):
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = max_length + 2
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
